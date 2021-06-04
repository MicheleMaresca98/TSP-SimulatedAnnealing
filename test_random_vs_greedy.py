import openpyxl
from TSP import TSP
from TSP_random import TSP_RANDOM
from leggi_coordinate import leggi_coordinate


if __name__ == "__main__":
    nuovo_file = openpyxl.Workbook()


    indice_linea_dimensione = 4
    indice_inizio_linee_coordinate = 7
    coordinate_a280, dimensione_a280 = leggi_coordinate("IstanzeBenchmark/a280.txt", indice_linea_dimensione, indice_inizio_linee_coordinate)
    coordinate_vm1084, dimensione_vm1084 = leggi_coordinate("IstanzeBenchmark/vm1084.tsp.txt", indice_linea_dimensione, indice_inizio_linee_coordinate)
    sheet = nuovo_file.create_sheet('remake_a280_vm1084_random')
    tsp_random_a280 = TSP_RANDOM(coordinate_a280)
    tsp_random_vm1084 = TSP_RANDOM(coordinate_vm1084)
    obj_finale_a280,obj_iniziale_a280 = tsp_random_a280.simulatedannealing()
    obj_finale_vm1084,obj_iniziale_vm1084 = tsp_random_vm1084.simulatedannealing()
    sheet['A1'] = 'remake'
    sheet['A2'] = 'L'
    sheet['A3'] = 'alpha'
    sheet['A4'] = 'temp_iniz'
    sheet['A5'] = 'temp_finale'
    sheet['A6'] = 'obj_sol_in'
    sheet['A7'] = 'obj_sol_fin'
    sheet['A8'] = 'best_solution'
    sheet['A9'] = 'gap'
    sheet['B1'] = 'a280'
    sheet['B2'] = tsp_random_a280.N
    sheet['B3'] = tsp_random_a280.alpha
    sheet['B4'] = tsp_random_a280.temperatura_iniziale
    sheet['B5'] = tsp_random_a280.temperatura_finale
    sheet['B6'] = obj_iniziale_a280
    sheet['B7'] = obj_finale_a280
    sheet['B8'] = 2579
    sheet['B9'] = (obj_finale_a280-2579)/2579
    sheet['C1'] = 'vm1084'
    sheet['C2'] = tsp_random_vm1084.N
    sheet['C3'] = tsp_random_vm1084.alpha
    sheet['C4'] = tsp_random_vm1084.temperatura_iniziale
    sheet['C5'] = tsp_random_vm1084.temperatura_finale
    sheet['C6'] = obj_iniziale_vm1084
    sheet['C7'] = obj_finale_vm1084
    sheet['C8'] = 239297
    sheet['C9'] = (obj_finale_vm1084-239297)/239297
    print('fine')
    nuovo_file.save('Test/Test_remake_a280_vm1084_random/remake_a280_vm1084_random.xlsx')

'''''''''''

    L = 280
    t_in = 100
    t_fin = 1
    alpha = 0.99
    indice_linea_dimensione = 4
    indice_inizio_linee_coordinate = 7
    coordinate_rl, dimensione_rl = leggi_coordinate("IstanzeBenchmark/rl5934.tsp.txt", indice_linea_dimensione, indice_inizio_linee_coordinate)
    coordinate_usa, dimensione_usa = leggi_coordinate("IstanzeBenchmark/usa13509.tsp.txt", 7,10)
    sheet = nuovo_file.create_sheet('instanze_grandi')
    tsp_rl = TSP(coordinate_rl,t_in,t_fin,alpha,L)
    tsp_usa = TSP(coordinate_usa,t_in,t_fin,alpha,L)
    obj_finale_rl = tsp_rl.simulatedannealing()
    print("rl5934")
    obj_finale_usa= tsp_usa.simulatedannealing()
    print("usa13509")
    sheet['A1'] = 'istanze_grandi'
    sheet['A2'] = 'L'
    sheet['A3'] = 'alpha'
    sheet['A4'] = 'temp_iniz'
    sheet['A5'] = 'temp_finale'
    sheet['A6'] = 'obj_sol_fin'
    sheet['A7'] = 'best_solution'
    sheet['A8'] = 'gap'
    sheet['B1'] = 'rl5934'
    sheet['B2'] = L
    sheet['B3'] = alpha
    sheet['B4'] = t_in
    sheet['B5'] = t_fin
    sheet['B6'] = obj_finale_rl
    sheet['B7'] = 554070
    sheet['B8'] = (obj_finale_rl-554070)/554070
    sheet['C1'] = 'usa13509'
    sheet['C2'] = L
    sheet['C3'] = alpha
    sheet['C4'] = t_in
    sheet['C5'] = t_fin
    sheet['C6'] = obj_finale_usa
    sheet['C7'] = 19947008
    sheet['C8'] = (obj_finale_usa-19947008)/19947008
    print('berlin52')

    indice_linea_dimensione = 4
    indice_inizio_linee_coordinate = 7
    coordinate, dimensione = leggi_coordinate("IstanzeBenchmark/a280.txt", indice_linea_dimensione, indice_inizio_linee_coordinate)
    sheet = nuovo_file.create_sheet('a280')
    tsp = TSP(coordinate)
    tsp_random = TSP_RANDOM(coordinate)
    obj_finale,obj_iniziale = tsp.simulatedannealing()
    obj_finale_random,obj_iniziale_random = tsp_random.simulatedannealing()
    sheet['A1'] = 'a280'
    sheet['A2'] = 'L'
    sheet['A3'] = 'alpha'
    sheet['A4'] = 'temp_iniz'
    sheet['A5'] = 'temp_finale'
    sheet['A6'] = 'obj_sol_in'
    sheet['A7'] = 'obj_sol_fin'
    sheet['B1'] = 'random'
    sheet['B2'] = tsp_random.N
    sheet['B3'] = tsp_random.alpha
    sheet['B4'] = tsp_random.temperatura_iniziale
    sheet['B5'] = tsp_random.temperatura_finale
    sheet['B6'] = obj_iniziale_random
    sheet['B7'] = obj_finale_random
    sheet['C1'] = 'greedy'
    sheet['C2'] = tsp.N
    sheet['C3'] = tsp.alpha
    sheet['C4'] = tsp.temperatura_iniziale
    sheet['C5'] = tsp.temperatura_finale
    sheet['C6'] = obj_iniziale
    sheet['C7'] = obj_finale
    print('a280')


    indice_linea_dimensione = 4
    indice_inizio_linee_coordinate = 7
    coordinate, dimensione = leggi_coordinate("IstanzeBenchmark/vm1084.tsp.txt", indice_linea_dimensione, indice_inizio_linee_coordinate)
    sheet = nuovo_file.create_sheet('vm1084')
    tsp = TSP(coordinate)
    tsp_random = TSP_RANDOM(coordinate)
    obj_finale,obj_iniziale = tsp.simulatedannealing()
    obj_finale_random,obj_iniziale_random = tsp_random.simulatedannealing()
    sheet['A1'] = 'vm1084'
    sheet['A2'] = 'L'
    sheet['A3'] = 'alpha'
    sheet['A4'] = 'temp_iniz'
    sheet['A5'] = 'temp_finale'
    sheet['A6'] = 'obj_sol_in'
    sheet['A7'] = 'obj_sol_fin'
    sheet['B1'] = 'random'
    sheet['B2'] = tsp_random.N
    sheet['B3'] = tsp_random.alpha
    sheet['B4'] = tsp_random.temperatura_iniziale
    sheet['B5'] = tsp_random.temperatura_finale
    sheet['B6'] = obj_iniziale_random
    sheet['B7'] = obj_finale_random
    sheet['C1'] = 'greedy'
    sheet['C2'] = tsp.N
    sheet['C3'] = tsp.alpha
    sheet['C4'] = tsp.temperatura_iniziale
    sheet['C5'] = tsp.temperatura_finale
    sheet['C6'] = obj_iniziale
    sheet['C7'] = obj_finale
    print('vm1084')

'''''
