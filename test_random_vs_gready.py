import openpyxl
from TSP import TSP
from TSP_random import TSP_RANDOM
from leggi_coordinate import leggi_coordinate


if __name__ == "__main__":
    nuovo_file = openpyxl.Workbook()
    indice_linea_dimensione = 4
    indice_inizio_linee_coordinate = 9
    coordinate, dimensione = leggi_coordinate("IstanzeBenchmark/burma14.txt", indice_linea_dimensione, indice_inizio_linee_coordinate)
    sheet = nuovo_file.create_sheet('burma')
    tsp = TSP(coordinate)
    tsp_random = TSP_RANDOM(coordinate)
    obj_finale,obj_iniziale = tsp.simulatedannealing()
    obj_finale_random,obj_iniziale_random = tsp_random.simulatedannealing()
    sheet['A1'] = 'burma14'
    sheet['A2'] = 'L'
    sheet['A3'] = 'alpha'
    sheet['A4'] = 'temp_iniz'
    sheet['A5'] = 'temp_finale'
    sheet['A6'] = 'obj_sol_in'
    sheet['A7'] = 'obj_sol_fin'
    sheet['B1'] = 'random'
    sheet['B2'] = tsp_random.N
    sheet['B3'] = tsp_random.alpha
    sheet['B4'] = tsp_random.temperatura_iniziale
    sheet['B5'] = tsp_random.temperatura_finale
    sheet['B6'] = obj_iniziale_random
    sheet['B7'] = obj_finale_random
    sheet['C1'] = 'greedy'
    sheet['C2'] = tsp.N
    sheet['C3'] = tsp.alpha
    sheet['C4'] = tsp.temperatura_iniziale
    sheet['C5'] = tsp.temperatura_finale
    sheet['C6'] = obj_iniziale
    sheet['C7'] = obj_finale
    print('burma14')


    indice_linea_dimensione = 4
    indice_inizio_linee_coordinate = 7
    coordinate, dimensione = leggi_coordinate("IstanzeBenchmark/berlin52.tsp", indice_linea_dimensione, indice_inizio_linee_coordinate)
    sheet = nuovo_file.create_sheet('berlin52')
    tsp = TSP(coordinate)
    tsp_random = TSP_RANDOM(coordinate)
    obj_finale,obj_iniziale = tsp.simulatedannealing()
    obj_finale_random,obj_iniziale_random = tsp_random.simulatedannealing()
    sheet['A1'] = 'berlin52'
    sheet['A2'] = 'L'
    sheet['A3'] = 'alpha'
    sheet['A4'] = 'temp_iniz'
    sheet['A5'] = 'temp_finale'
    sheet['A6'] = 'obj_sol_in'
    sheet['A7'] = 'obj_sol_fin'
    sheet['B1'] = 'random'
    sheet['B2'] = tsp_random.N
    sheet['B3'] = tsp_random.alpha
    sheet['B4'] = tsp_random.temperatura_iniziale
    sheet['B5'] = tsp_random.temperatura_finale
    sheet['B6'] = obj_iniziale_random
    sheet['B7'] = obj_finale_random
    sheet['C1'] = 'greedy'
    sheet['C2'] = tsp.N
    sheet['C3'] = tsp.alpha
    sheet['C4'] = tsp.temperatura_iniziale
    sheet['C5'] = tsp.temperatura_finale
    sheet['C6'] = obj_iniziale
    sheet['C7'] = obj_finale
    print('berlin52')

    indice_linea_dimensione = 4
    indice_inizio_linee_coordinate = 7
    coordinate, dimensione = leggi_coordinate("IstanzeBenchmark/a280.txt", indice_linea_dimensione, indice_inizio_linee_coordinate)
    sheet = nuovo_file.create_sheet('a280')
    tsp = TSP(coordinate)
    tsp_random = TSP_RANDOM(coordinate)
    obj_finale,obj_iniziale = tsp.simulatedannealing()
    obj_finale_random,obj_iniziale_random = tsp_random.simulatedannealing()
    sheet['A1'] = 'a280'
    sheet['A2'] = 'L'
    sheet['A3'] = 'alpha'
    sheet['A4'] = 'temp_iniz'
    sheet['A5'] = 'temp_finale'
    sheet['A6'] = 'obj_sol_in'
    sheet['A7'] = 'obj_sol_fin'
    sheet['B1'] = 'random'
    sheet['B2'] = tsp_random.N
    sheet['B3'] = tsp_random.alpha
    sheet['B4'] = tsp_random.temperatura_iniziale
    sheet['B5'] = tsp_random.temperatura_finale
    sheet['B6'] = obj_iniziale_random
    sheet['B7'] = obj_finale_random
    sheet['C1'] = 'greedy'
    sheet['C2'] = tsp.N
    sheet['C3'] = tsp.alpha
    sheet['C4'] = tsp.temperatura_iniziale
    sheet['C5'] = tsp.temperatura_finale
    sheet['C6'] = obj_iniziale
    sheet['C7'] = obj_finale
    print('a280')


    indice_linea_dimensione = 4
    indice_inizio_linee_coordinate = 7
    coordinate, dimensione = leggi_coordinate("IstanzeBenchmark/vm1084.tsp.txt", indice_linea_dimensione, indice_inizio_linee_coordinate)
    sheet = nuovo_file.create_sheet('vm1084')
    tsp = TSP(coordinate)
    tsp_random = TSP_RANDOM(coordinate)
    obj_finale,obj_iniziale = tsp.simulatedannealing()
    obj_finale_random,obj_iniziale_random = tsp_random.simulatedannealing()
    sheet['A1'] = 'vm1084'
    sheet['A2'] = 'L'
    sheet['A3'] = 'alpha'
    sheet['A4'] = 'temp_iniz'
    sheet['A5'] = 'temp_finale'
    sheet['A6'] = 'obj_sol_in'
    sheet['A7'] = 'obj_sol_fin'
    sheet['B1'] = 'random'
    sheet['B2'] = tsp_random.N
    sheet['B3'] = tsp_random.alpha
    sheet['B4'] = tsp_random.temperatura_iniziale
    sheet['B5'] = tsp_random.temperatura_finale
    sheet['B6'] = obj_iniziale_random
    sheet['B7'] = obj_finale_random
    sheet['C1'] = 'greedy'
    sheet['C2'] = tsp.N
    sheet['C3'] = tsp.alpha
    sheet['C4'] = tsp.temperatura_iniziale
    sheet['C5'] = tsp.temperatura_finale
    sheet['C6'] = obj_iniziale
    sheet['C7'] = obj_finale
    print('vm1084')

    nuovo_file.save('Test/Random_VS_Gready_standard_parameter/Random_vs_Gready.xlsx')


